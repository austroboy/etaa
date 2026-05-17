"""
CV Module – Services.

CV collection, text extraction, LLM-driven scoring, and output packaging.

CHANGED from original:
- package_results() now saves a standalone Excel file only (no ZIP).
- No duplicate Excel inside ZIP.
- Returns the xlsx path instead of a zip path.
- upsert_candidate_profile() added: creates or updates CandidateProfile
  (dedup by email OR phone).
"""

import csv
import io
import json
import logging
import os
import re
from pathlib import Path
from typing import List

from django.conf import settings

from apps.llm_client import get_llm_client

logger = logging.getLogger("etaa")

from apps.cv_module.models import CandidateProfile

# ── CV Collection ─────────────────────────────────────────────────────────────


def collect_cvs_from_local(directory: str) -> List[str]:
    """Return list of CV file paths from a local directory."""
    path = Path(directory)
    if not path.is_dir():
        raise FileNotFoundError(f"Directory not found: {directory}")
    extensions = {".pdf", ".docx", ".doc"}
    return [str(f) for f in path.rglob("*") if f.suffix.lower() in extensions]


def collect_cvs_from_drive(folder_link: str, dest_dir: str) -> List[str]:
    """Download CV files from a Google Drive folder link to dest_dir."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
        from google.oauth2 import service_account

        creds_json = settings.GOOGLE_SERVICE_ACCOUNT_JSON
        if not creds_json:
            raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON not configured")

        creds_info = json.loads(creds_json)
        credentials = service_account.Credentials.from_service_account_info(
            creds_info,
            scopes=["https://www.googleapis.com/auth/drive.readonly"],
        )
        service = build("drive", "v3", credentials=credentials)

        if "folders/" in folder_link:
            folder_id = folder_link.split("folders/")[1].split("?")[0]
        else:
            folder_id = folder_link

        query = f"'{folder_id}' in parents and trashed=false"
        results = service.files().list(
            q=query, fields="files(id, name, mimeType)").execute()
        files = results.get("files", [])

        os.makedirs(dest_dir, exist_ok=True)
        downloaded = []
        allowed_mime = {
            "application/pdf",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        }

        for file in files:
            if file["mimeType"] not in allowed_mime:
                continue
            dest_path = os.path.join(dest_dir, file["name"])
            request = service.files().get_media(fileId=file["id"])
            with open(dest_path, "wb") as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
            downloaded.append(dest_path)
            logger.info("Downloaded CV: %s", file["name"])

        return downloaded

    except Exception as exc:  # noqa: BLE001
        logger.error("Google Drive CV collection failed: %s", exc)
        raise


# ── Text Extraction ───────────────────────────────────────────────────────────


def extract_text_from_cv(file_path: str) -> str:
    """Extract plain text from a PDF or DOCX CV file."""
    ext = Path(file_path).suffix.lower()
    try:
        if ext == ".pdf":
            return _extract_pdf(file_path)
        if ext in {".docx", ".doc"}:
            return _extract_docx(file_path)
        logger.warning("Unsupported CV format: %s", file_path)
        return ""
    except Exception as exc:  # noqa: BLE001
        logger.warning("Text extraction failed for %s: %s", file_path, exc)
        return ""


def _extract_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        return "\n".join(
            page.extract_text() or "" for page in reader.pages
        )
    except Exception:  # noqa: BLE001
        try:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                return "\n".join(
                    page.extract_text() or "" for page in pdf.pages
                )
        except Exception as exc2:  # noqa: BLE001
            logger.warning("PDF extraction failed (%s): %s", path, exc2)
            return ""


def _extract_docx(path: str) -> str:
    try:
        import docx2txt
        return docx2txt.process(path) or ""
    except Exception:  # noqa: BLE001
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as exc2:  # noqa: BLE001
            logger.warning("DOCX extraction failed (%s): %s", path, exc2)
            return ""


# ── LLM Scoring ───────────────────────────────────────────────────────────────

SCORE_SYSTEM = """
You are an expert HR assistant. Analyse the CV and score it against the job requirements.
Return ONLY a JSON object with these exact keys:
{
  "candidate_name": "",
  "current_designation": "",
  "current_company": "",
  "previous_designation": "",
  "previous_company": "",
  "years_experience": 0,
  "relevant_industries": "",
  "email": "",
  "phone": "",
  "location": "",
  "academic_qualification": "",
  "match_score": 0,
  "key_qualifications": "",
  "summary": ""
}
Rules:
- match_score: integer 0-100 reflecting fit for the role.
- years_experience: total professional years as integer.
- academic_qualification: highest degree + institution.
- key_qualifications: 3-5 bullet points (plain text, no markdown).
- summary: 2-3 sentence narrative.
- Return ONLY the JSON object, no prose.
""".strip()


def _safe_int(val) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0


def _extract_json_object(s: str) -> dict:
    """Find and parse the first {...} JSON object in a string."""
    depth, start, in_string = 0, None, False
    for i, ch in enumerate(s):
        if ch == '"' and (i == 0 or s[i - 1] != "\\"):
            in_string = not in_string
        if not in_string:
            if ch == "{" and depth == 0:
                start = i
                in_string = False
                depth += 1
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    candidate = s[start:i + 1]
                    return json.loads(candidate)
    raise ValueError(f"unbalanced JSON in response: {s[:200]}")


def _heuristic_name_from_text(cv_text: str) -> str:
    """Fallback: try to read a likely name from the first non-empty line."""
    for line in (cv_text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if "@" in line or "http" in line.lower():
            continue
        if re.search(r"\d{4,}", line):
            continue
        if line.lower() in {"curriculum vitae", "cv", "resume", "résumé"}:
            continue
        words = line.split()
        if 1 < len(words) <= 5 and all(
            w[:1].isupper() or not w[:1].isalpha() for w in words
        ):
            return line[:80]
        break
    return ""


_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_PHONE_RE = re.compile(r"\+?\d[\d\s().\-]{7,}\d")


def score_cv(cv_text: str, job_requirements: str, file_name: str = "") -> dict:
    """Score a single CV against job requirements using the LLM."""
    defaults = {
        "candidate_name":       "",
        "current_designation":  "",
        "current_company":      "",
        "previous_designation": "N/A",
        "previous_company":     "N/A",
        "years_experience":     0,
        "relevant_industries":  "",
        "email":                "",
        "phone":                "",
        "location":             "",
        "match_score":          0,
        "key_qualifications":   "",
        "summary":              "",
        "academic_qualification": "N/A",
    }

    if not cv_text or len(cv_text.strip()) < 30:
        logger.warning("CV %s has no extractable text", file_name)
        defaults["summary"] = "Could not extract text from CV file"
        return defaults

    llm = get_llm_client()

    # Cost-optimised CV scoring path:
    #   * Use Haiku 4.5 — 3x cheaper input/output than Sonnet 4.5,
    #     SWE-bench accuracy gap is ~5% on hard reasoning; for CV
    #     extraction (a structured-output task) the gap is much smaller.
    #   * Cap CV text at 4000 chars (~1000 tokens). Most CVs have
    #     all decision-relevant info in the first 1-2 pages.
    #   * Mark the system prompt as cacheable. After the first call in
    #     a batch, subsequent calls within ~5 min pay 90% less for the
    #     system tokens. With ~600-token system prompts and 100s of CVs,
    #     this saves ~$0.0005 per CV.
    from apps.llm_client import LLMClient as _LC
    cv_text_capped = cv_text[:4000]
    prompt = (
        f"JOB REQUIREMENTS:\n{job_requirements}\n\n"
        f"CV TEXT:\n{cv_text_capped}"
    )
    raw = ""
    try:
        raw = llm.complete(
            prompt, system=SCORE_SYSTEM,
            max_tokens=1200, temperature=0.0,
            model=_LC.ANTHROPIC_HAIKU_MODEL,
            cache_system=True,
        )
        data = _extract_json_object(raw)
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "CV scoring LLM/parse error for %s: %s | raw=%r",
            file_name, exc, (raw or "")[:300],
        )
        data = {}

    out = dict(defaults)
    for k, v in (data or {}).items():
        if k in out:
            out[k] = v

    out["match_score"]      = max(0, min(100, _safe_int(out.get("match_score"))))
    out["years_experience"] = max(0, _safe_int(out.get("years_experience")))
    out["candidate_name"]   = (out.get("candidate_name") or "").strip()
    out["email"]            = (out.get("email") or "").strip()
    out["phone"]            = (out.get("phone") or "").strip()

    if not out["candidate_name"] or out["candidate_name"].lower() == "unknown":
        guess = _heuristic_name_from_text(cv_text)
        if guess:
            out["candidate_name"] = guess

    if not out["email"]:
        m = _EMAIL_RE.search(cv_text)
        if m:
            out["email"] = m.group(0)

    if not out["phone"]:
        m = _PHONE_RE.search(cv_text)
        if m:
            out["phone"] = m.group(0).strip()

    if not out["candidate_name"]:
        stem = os.path.splitext(file_name or "")[0]
        if stem:
            out["candidate_name"] = stem.replace("_", " ").replace("-", " ")[:80]

    return out


def rank_all_cvs(file_paths: List[str], job_requirements: str) -> List[dict]:
    """Extract, score, and rank all CVs. Returns sorted list (highest score first)."""
    results = []
    for fp in file_paths:
        text = extract_text_from_cv(fp)
        score_data = score_cv(text, job_requirements, file_name=os.path.basename(fp))
        results.append({
            "file_path":              fp,
            "file_name":              os.path.basename(fp),
            "candidate_name":         score_data["candidate_name"] or "(name not detected)",
            "current_designation":    score_data["current_designation"],
            "current_company":        score_data["current_company"],
            "previous_designation":   score_data.get("previous_designation", "N/A"),
            "previous_company":       score_data.get("previous_company", "N/A"),
            "years_experience":       score_data["years_experience"],
            "relevant_industries":    score_data["relevant_industries"],
            "email":                  score_data["email"],
            "phone":                  score_data["phone"],
            "location":               score_data["location"],
            "academic_qualification": score_data.get("academic_qualification", "N/A"),
            "match_score":            score_data["match_score"],
            "key_qualifications":     score_data["key_qualifications"],
            "summary":                score_data["summary"],
        })

    results.sort(key=lambda x: x["match_score"], reverse=True)
    for rank, r in enumerate(results, start=1):
        r["rank"] = rank
    return results


# ── Candidate Profile Upsert (dedup by email OR phone) ───────────────────────


def upsert_candidate_profile(entry: dict, job) -> "CandidateProfile":  # noqa: F821
    """
    Create or update a CandidateProfile for this candidate.

    Matching priority:
      1. email (if non-empty)
      2. phone (if non-empty)
    If neither matches an existing profile, a new one is created.
    HR annotation fields (remark, rejected_company, others, is_active)
    are NEVER overwritten by this function.
    """
    from apps.cv_module.models import CandidateProfile

    email = (entry.get("email") or "").strip()
    phone = (entry.get("phone") or "").strip()

    profile = None

    # Try to find existing profile
    if email:
        profile = CandidateProfile.objects.filter(email=email).first()
    if profile is None and phone:
        profile = CandidateProfile.objects.filter(phone=phone).first()

    cv_fields = {
        "candidate_name":         entry.get("candidate_name", ""),
        "current_designation":    entry.get("current_designation", ""),
        "current_company":        entry.get("current_company", ""),
        "previous_designation":   entry.get("previous_designation", "N/A"),
        "previous_company":       entry.get("previous_company", "N/A"),
        "years_experience":       entry.get("years_experience", 0),
        "relevant_industries":    entry.get("relevant_industries", ""),
        "location":               entry.get("location", ""),
        "academic_qualification": entry.get("academic_qualification", "N/A"),
        "key_qualifications":     entry.get("key_qualifications", ""),
        "summary":                entry.get("summary", ""),
        "match_score":            entry.get("match_score", 0),
        "rank":                   entry.get("rank", 0),
        "file_name":              entry.get("file_name", ""),
        "email":                  email,
        "phone":                  phone,
        "latest_job":             job,
    }

    if profile is None:
        profile = CandidateProfile.objects.create(**cv_fields)
        logger.info("Created CandidateProfile id=%s for %s", profile.pk, email or phone)
    else:
        for field, value in cv_fields.items():
            setattr(profile, field, value)
        profile.save()
        logger.info("Updated CandidateProfile id=%s for %s", profile.pk, email or phone)

    return profile


# ── Output Packaging ──────────────────────────────────────────────────────────

# Columns in the Excel scoring sheet, in display order.
_EXCEL_COLUMNS = [
    ("rank",                    "Rank"),
    ("candidate_name",          "Candidate Name"),
    ("match_score",             "Score"),
    ("academic_qualification",  "Academic Qualification"),
    ("current_designation",     "Current Designation"),
    ("previous_designation",    "Previous Designation"),
    ("previous_company",        "Previous Company"),
    ("current_company",         "Current Company"),
    ("years_experience",        "Experience (yrs)"),
    ("relevant_industries",     "Relevant Industries"),
    ("location",                "Location"),
    ("email",                   "Email"),
    ("phone",                   "Phone"),
    ("key_qualifications",      "Key Qualifications"),
    ("summary",                 "Summary"),
    ("file_name",               "File"),
]


def _build_excel_summary(ranked: List[dict], job_requirements: str,
                         out_path: str) -> str:
    """Write a styled .xlsx scoring summary. Returns the file path."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed; skipping Excel output")
        return ""

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Rankings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496",
                              fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, (_, header) in enumerate(_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(ranked, start=2):
        for col_idx, (key, _) in enumerate(_EXCEL_COLUMNS, start=1):
            val = r.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap

    widths = {
        "Rank": 6, "Candidate Name": 24, "Score": 7,
        "Current Designation": 26, "Current Company": 24,
        "Experience (yrs)": 10, "Relevant Industries": 28,
        "Location": 18, "Email": 30, "Phone": 18,
        "Key Qualifications": 40, "Summary": 50, "File": 28,
        "Academic Qualification": 30,
        "Previous Designation": 26,
        "Previous Company": 24,
    }
    for col_idx, (_, header) in enumerate(_EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 18)

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Job Specification")
    ws2["A1"] = "Job Requirements"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = job_requirements
    ws2["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 100
    ws2.row_dimensions[3].height = 300

    wb.save(out_path)
    return out_path


def package_results(ranked: List[dict], top_n: int, output_dir: str,
                    job_id: int, job_requirements: str = "") -> str:
    """
    Save the Excel scoring summary to output_dir.

    CHANGED: No longer creates a ZIP file. Returns the xlsx path.
    The top_n parameter is kept for API compatibility but only affects
    how the file is named (for clarity).
    """
    os.makedirs(output_dir, exist_ok=True)
    actual_top = min(top_n, len(ranked))

    xlsx_path = os.path.join(
        output_dir, f"CV_Ranking_Summary_Job{job_id}.xlsx",
    )
    result = _build_excel_summary(ranked, job_requirements, xlsx_path)
    if result:
        logger.info(
            "Saved Excel summary for job %s (%d CVs ranked, top %d requested) → %s",
            job_id, len(ranked), actual_top, xlsx_path,
        )
    return xlsx_path